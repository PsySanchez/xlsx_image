import os
from openpyxl import load_workbook, Workbook
from PIL import Image

# Specify valid image extensions
valid_image_extensions = (".jpg", ".jpeg", ".png", ".webp")


def process_images(image_folder, output_folder, barcode_column="barcode"):
    # Locate the first Excel file in the current directory or set to None if not found
    xlsx_file = next((file for file in os.listdir() if file.endswith(".xlsx")), None)
    worksheet = None
    dynamic_output_folder = None

    if xlsx_file:
        # Load the Excel workbook and active sheet
        workbook = load_workbook(xlsx_file)
        worksheet = workbook.active
        # Get the header row and locate the barcode column
        header = next(worksheet.iter_rows(max_row=1, values_only=True))
        try:
            barcode_index = header.index(barcode_column)
        except ValueError:
            raise ValueError(f"Column '{barcode_column}' not found in the Excel file.")
    else:
        print("No Excel file found. Proceeding without barcode matching.")

    # Create the output folder if it doesn't exist
    os.makedirs(output_folder, exist_ok=True)

    # get list of sub fodlers in root folder
    folders = [f.path for f in os.scandir(image_folder) if f.is_dir()]
    folders.append(image_folder)

    for folder in folders:
        print(f"Processing directory: {folder}")

        # Continue if folder is empty from images
        if not any(
            file.lower().endswith(valid_image_extensions) for file in os.listdir(folder)
        ):
            print(f"Folder {folder} is empty of valid images. Skipping...")
            continue

        # If folder has different name with root folder, create output folder with same name in output folder
        if folder != image_folder:
            dynamic_output_folder = os.path.join(
                output_folder, os.path.basename(folder)
            )
            os.makedirs(dynamic_output_folder, exist_ok=True)
        else:
            dynamic_output_folder = output_folder

        if worksheet:
            # Create a 'not_found' folder to store images without a matching barcode
            not_found_folder = os.path.join(dynamic_output_folder, "not_found")
            os.makedirs(not_found_folder, exist_ok=True)

        # Prepare a list to track "not found" images
        not_found_images = []

        # Process each image in the image folder
        for filename in os.listdir(folder):
            if filename.lower().endswith(valid_image_extensions):
                img_path = os.path.join(folder, filename)
                file_basename = filename.rsplit(".", 1)[0]  # Remove file extension

                found = False
                if worksheet:
                    # Check if the image filename exists in the Excel file
                    for row in worksheet.iter_rows(values_only=True):
                        if any(
                            file_basename
                            == (
                                str(int(cell))
                                if isinstance(cell, (int, float)) and cell == int(cell)
                                else str(cell)
                            )
                            for cell in row
                        ):
                            cell = row[barcode_index]
                            barcode = (
                                str(int(cell))
                                if isinstance(cell, (int, float)) and cell == int(cell)
                                else str(cell)
                            )
                            found = True
                            break

                try:
                    # Open the image and resize it
                    with Image.open(img_path) as img:
                        # Force loading of the image
                        img.load()

                        # Check if the image is in P mode with transparency
                        if img.mode == "P" and "transparency" in img.info:
                            # Convert to RGBA first if transparency exists
                            img = img.convert("RGBA")
                        elif img.mode == "P":
                            # Convert to RGB if no transparency exists
                            img = img.convert("RGB")

                        # Convert RGBA to RGB for JPEG compatibility
                        if img.mode == "RGBA":
                            img = img.convert("RGB")

                        # Convert to RGB if no transparency exists
                        img = img.resize((270, 300))

                except Exception as e:
                    print(f"Error processing image {filename}: {e}")
                    continue

                if found:
                    output_filename = f"{barcode}.jpg"
                    dynamic_output_folder = output_folder

                elif worksheet:
                    output_filename = f"{file_basename}.jpg"
                    dynamic_output_folder = not_found_folder
                    not_found_images.append(file_basename)
                    print(
                        f"Barcode not found for image {filename}. Saved as {output_filename}"
                    )
                else:
                    output_filename = filename

                # Save the image
                img.save(os.path.join(dynamic_output_folder, output_filename), "JPEG")

        # Create an Excel file for "not found" images
        if worksheet:
            not_found_excel_path = os.path.join(
                not_found_folder, "not_found_images.xlsx"
            )
            create_not_found_excel(not_found_images, not_found_excel_path)


def create_not_found_excel(image_names, output_excel_path):
    """Creates an Excel file listing all the images in the 'not_found' folder."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Not Found Images"
    sheet.append(["Image Filename"])  # Add header

    # Add image filenames to the sheet
    for image_name in image_names:
        sheet.append([image_name])

    # Save the Excel file
    workbook.save(output_excel_path)
    print(f"Excel file for 'not found' images created: {output_excel_path}")


# Example usage
image_folder = "./images"
output_folder = "output_images"
process_images(image_folder, output_folder)
